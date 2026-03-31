from openpyxl import load_workbook
import os
from os.path import join, dirname
from dotenv import load_dotenv
import requests
import re
import html

def clean_description(raw):
    """Nettoie la description (HTML ou texte brut)."""
    if not raw:
        return ""
    text = re.sub(r'<br\s*/?>|</p>|</div>|</li>', '\n', raw, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()

def build_updated_description(description, email, text_to_add):
    clean = clean_description(description)
    if not clean:
        return None

    blocks = re.split(r'(\s*-{5,}\s*)', clean)
    updated = False

    for i in range(0, len(blocks), 2):
        block = blocks[i]

        email_match = re.search(r"Email\s*:\s*(.+)", block)
        if not email_match:
            continue

        found_email = email_match.group(1).strip()
        if found_email.lower() != email.lower():
            continue

        source_match = re.search(r"Source\s*:\s*(.+)", block)

        if source_match:
            old_source = source_match.group(1).strip()
            new_source = old_source + ", " + text_to_add

            def repl(match):
                return match.group(1) + new_source

            block = re.sub(r"(Source\s*:\s*)(.+)", repl, block)
        else:
            block = block.rstrip() + f"\nSource : {text_to_add}\n"

        blocks[i] = block
        updated = True
        break

    if not updated:
        return None

    return "".join(blocks)





# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────

dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

NOCRM_API_KEY = os.environ.get("NOCRM_API_KEY")
NOCRM_SUBDOMAIN = os.environ.get("NOCRM_SUBDOMAIN")

if not NOCRM_API_KEY or not NOCRM_SUBDOMAIN:
    raise ValueError("Variables d'environnement manquantes (NOCRM_API_KEY, NOCRM_SUBDOMAIN)")

BASE_URL = f"https://{NOCRM_SUBDOMAIN}.nocrm.io/api/v2"
HEADERS = {
    "X-API-KEY": NOCRM_API_KEY,
    "Content-Type": "application/json"
}

data_repository = "/webinar_to_crm/data_xlsx"
data_sheet_name = os.listdir(data_repository)

if not data_sheet_name : 
    print("=> Aucun fichier trouvé...")
else : 
    print("=> Fichier trouvé : ", data_sheet_name[0])

file_path = os.path.join(data_repository, data_sheet_name[0])

wb = load_workbook(file_path)
ws1 = wb.active

# ──────────────────────────────────────────────
# ──────────────────────────────────────────────
# ──────────────────────────────────────────────


column_to_delete = ["Code d'accès au webinar", "Lien d'accès", "Ville (depuis IP)", "Pays (depuis IP)",
                    "Ip country name", "Date d'inscription", "Durée de la présence (en seconde)", "Nombre de sessions inscrites",
                    "Nombre de sessions participées", "Sessions inscrites", "Sessions participées", "A vu le replay", 
                    "Inscrit à la demande", "Nombre de messages", "Nombre de questions", "Nb de sondages répondus", "Nombre d'upvotes questions", 
                    "Referrer", "Source (UTM)", "Médium (UTM)", "Term (UTM)", "Content (UTM)", "Campaign (UTM)", "Version du navigateur",
                    "Navigateur", "Système d'exploitation", "Version du système d'exploitation", "Hauteur de l'écran", "Largeur de l'écran",
                    "Is guest speaker", "Is team member"]

increment_deleting = 0

header_snapshot = [(cell.col_idx, cell.value) for cell in ws1[1]]

for col_idx, value in header_snapshot:
    if value in column_to_delete:
        ws1.delete_cols(col_idx - increment_deleting)
        increment_deleting += 1
print("=> Colonnes inutiles supprimées")

leads = []
for row in ws1.iter_rows(min_row=2, max_row=4, values_only=True):
    lead = {
        "Email": row[0] or "",
        "Prénom": row[1] or "",
        "Nom": row[2] or "",
        "Présent": row[3] or "",
        "Taux de présence": row[4] or "",
        "Company": row[5] or "",
        "Job title": row[6] or "",
        "Phone number": row[7] or ""
    }
    leads.append(lead)

print(leads)

offset = 0
limit = 100
api_calls = 0

for lead in leads:
    params = {"limit": limit, "offset": offset, "email": lead["Email"]}
    response = requests.get(f"{BASE_URL}/leads", headers=HEADERS, params=params)
    api_calls += 1

    if response.status_code != 200:
        print(f"❌ Erreur API: {response.status_code} - {response.text}")
        continue

    api_leads = response.json()

    if not api_leads:
        print("=> Call API :", api_calls)
        print("=> Pas de lead connu pour l'adresse email :", lead["Email"])
        continue

    api_lead = api_leads[0]

    lead_title = api_lead.get("title", "")
    lead_id = api_lead.get("id", "")
    lead_description = api_lead.get("description", "")

    prénom = lead["Prénom"]
    nom = lead["Nom"]
    présence = str(lead["Présent"]).lower()
    taux_présence = lead["Taux de présence"]

    if présence == "false":
        text_to_add = f"{prénom} {nom} s'est inscrit mais n'a pas participé au Webinar xxxxxxx"
    elif présence == "true":
        text_to_add = f"{prénom} {nom} a participé au Webinar xxxxxxx avec un taux de présence de {taux_présence}"
    else:
        text_to_add = f"{prénom} {nom} s'est inscrit au Webinar xxxxxx"

    new_description = build_updated_description(lead_description, lead["Email"], text_to_add)

    if not new_description:
        print("=> Aucun bloc contact trouvé pour :", lead["Email"])
        continue

    payload = {
        "description": new_description
    }

    response = requests.put(
        f"{BASE_URL}/leads/{lead_id}",
        headers=HEADERS,
        json=payload
    )
    api_calls += 1

    if response.status_code not in [200, 201, 204]:
        print(f"❌ Erreur PUT: {response.status_code} - {response.text}")
        continue

    print("=> Call API :", api_calls)
    print(f"=> Email : {lead['Email']} --- Lead : {lead_title} --- ID : {lead_id}")

wb.save("data_cleaned/data_cleaned.xlsx")
print("=> Document enregistré")
